#ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡTermProject 2020105726 소프트웨어융합학과 이준혁ㅡㅡㅡㅡㅡㅡㅡㅡㅡ
#ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ엑셀기반 고객관리 프로그램ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
#ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ사용 라이브러리 : openpyxl, os ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ


import openpyxl
import os

# 기본정보사항 출력함수

def standard(a):
    while True:
        print("현재 " + a + "님의 기본정보사항 현황입니다. \n")
        try:
            print("이름 : " + ws['B5'].value + '\n')
        except:
            print("이름 : \n")
        try:
            print("연락처 : " + ws['E5'].value +'\n')
        except:
            print("연락처 : \n")
        try:
            print("주민등록번호 : "+ ws['B6'].value + '\n')
        except:
            print("주민등록번호 : \n")
        try:
            print("우편번호 : " + ws['B7'].value + '\n')
        except:
            print("우편번호 : \n")
        try:
            print("주소 : "+ ws['B8'].value + '\n')
        except:
            print("주소 : \n")
        try:
            print("E-mail : "+ ws['E7'].value +'\n')
        except:
            print("E-mail : \n")
        break

#직장정보사항 출력함수

def company(a):
    while True:
        print("현재" + a + "님의 직장정보사항입니다. \n")
        try:
            print("직장명 : " + ws['B10'].value +'\n')
        except:
            print("직장명 : \n")
        try:
            print("직장 주소 : " + ws['B11'].value + '\n')
        except:
            print("직장 주소 : \n")
        try:
            print("직장 연락처 : " + ws['E10'].value + '\n')
        except:
            print("직장 연락처 : \n")
        try:
            print("사업자 번호 : "+ ws['B12'].value + '\n')
        except:
            print("사업자 번호 : \n")
        try:
            print("Fax : " + ws['E12'].value + '\n')
        except:
            print("Fax : \n")
        break

#가족관계사항 출력함수

def family(a):
    while True:
        print("현재" + a + "님의 가족관계사항 현황입니다. \n")
        try:
            print("1 : " + ws['H6'].value + ws['I6'].value + ws['J6'].value + ws['K6'].value + '\n')
        except:
            print("1: \n")
        try:
            print("2 : " + ws['H7'].value + ws['I7'].value + ws['J7'].value + ws['K7'].value + '\n')
        except:
            print("2: \n")
        try:
            print("3 : " + ws['H8'].value + ws['I8'].value + ws['J8'].value + ws['K8'].value + '\n')
        except:
            print("3: \n")
        try:
            print("4 : " + ws['H9'].value + ws['I9'].value + ws['J9'].value + ws['K9'].value + '\n')
        except:
            print("4: \n")
        try:
            print("5 : " + ws['H10'].value + ws['I10'].value + ws['J10'].value + ws['K10'].value + '\n')
        except:
            print("5: \n")
        try:
            print("6 : " + ws['H11'].value + ws['I11'].value + ws['J11'].value + ws['K11'].value + '\n')
        except:
            print("6: \n")
        break

#자동차 보유현황 출력함수

def car(a):
    while True:
        print("현재"  + a + "님의 자동차 보유 현황입니다. \n")
        try:
            print("1 : " + ws['A15'].value + ws['C15'].value + ws['D15'].value + ws['F15'].value + '\n')
        except:
            print("1: \n")
        try:
            print("2 : " + ws['A16'].value + ws['C16'].value + ws['D16'].value + ws['F16'].value + '\n')
        except:
            print("2: \n")
        try:
            print("3 : " + ws['A17'].value + ws['C17'].value + ws['D17'].value + ws['F17'].value + '\n')
        except:
            print("3: \n")
        try:
            print("4 : " + ws['A18'].value + ws['C18'].value + ws['D18'].value + ws['F18'].value + '\n')
        except:
            print("4: \n")
        try:
            print("5 : " + ws['A19'].value + ws['C19'].value + ws['D19'].value + ws['F19'].value + '\n')
        except:
            print("5: \n")
        try:
            print("6 : " + ws['A20'].value + ws['C20'].value + ws['D20'].value + ws['F20'].value + '\n')
        except:
            print("6: \n")
        break

#자사보험 가입현황 및 상태 출력함수

def insurance(a):
    while True:
        print("현재" + a + "님의 자사보험 가입현황 및 상태입니다. \n")
        try:
            print("1 : " + ws['A23'].value + ws['B23'].value + ws['C23'].value + ws['E23'].value + ws['H23'].value + ws['J23'].value + ws['K23'].value + '\n')
        except:
            print("1 : \n")
        try:
            print("2 : " + ws['A24'].value + ws['B24'].value + ws['C24'].value + ws['E24'].value + ws['H24'].value + ws['J24'].value + ws['K24'].value + '\n')
        except:
            print("2 : \n")
        try:
            print("3 : " + ws['A25'].value + ws['B25'].value + ws['C25'].value + ws['E25'].value + ws['H25'].value + ws['J25'].value + ws['K25'].value + '\n')
        except:
            print("3 : \n")
        try:
            print("4 : " + ws['A26'].value + ws['B26'].value + ws['C26'].value + ws['E26'].value + ws['H26'].value + ws['J26'].value + ws['K26'].value + '\n')
        except:
            print("4 : \n")
        try:
            print("5 : " + ws['A27'].value + ws['B27'].value + ws['C27'].value + ws['E27'].value + ws['H27'].value + ws['J27'].value + ws['K27'].value + '\n')
        except:
            print("5 : \n")
        try:
            print("6 : " + ws['A28'].value + ws['B28'].value + ws['C28'].value + ws['E28'].value + ws['H28'].value + ws['J28'].value + ws['K28'].value + '\n')
        except:
            print("6 : \n")
        try:
            print("7 : " + ws['A29'].value + ws['B29'].value + ws['C29'].value + ws['E29'].value + ws['H29'].value + ws['J29'].value + ws['K29'].value + '\n')
        except:
            print("7 : \n")
        break      

#특이사항 및 참고사항 출력함수

def special(a):
    while True:
        print("현재" + a + "님의 특이사항 및 참고사항 입니다. \n")
        try:
            print("1 : " + ws['A32'].value + '\n')
        except:
            print("1 : \n")
        try:
            print("2 : " + ws['A33'].value + '\n')
        except:
            print("2 : \n")
        try:
            print("3 : " + ws['A34'].value + '\n')
        except:
            print("3 : \n")
        try:
            print("4 : " + ws['A35'].value + '\n')
        except:
            print("4 : \n")
        try:
            print("5 : " + ws['A36'].value + '\n')
        except:
            print("5 : \n")
        try:
            print("6 : " + ws['A37'].value + '\n')
        except:
            print("6 : \n")
        try:
            print("7 : " + ws['A38'].value + '\n')
        except:
            print("7 : \n")
        try:
            print("8 : " + ws['A39'].value + '\n')
        except:
            print("8 : \n")
        break

#고객정보 출력함수

def find_info(a):
    standard(a)
    company(a)
    family(a)
    car(a)
    insurance(a)
    special(a)  

while True:
    print("**********************************주의 사항********************************")
    print("*   1. 반드시 고객카드.xlsx와 고객명.xlsx들을 같은 디렉토리에 저장하십시오.   *")
    print("*   2. 엑셀파일을 열어둔채로 프로그램을 실행시키지 마십시오.                  *")
    print("**************************************************************************")
    print("*                                Menu                                    *")
    print("**************************************************************************")
    print("*                           1.고객정보추가                                *")
    print("*                           2.고객정보수정                                *")
    print("*                           3.고객정보조회                                *")
    print("*                           4.고객정보삭제                                *")
    print("*                           0.   종료                                    *")
    print("**************************************************************************")
    doing = input("0,1,2,3,4 중 하나를 입력하세요.")
    try:
        if int(doing) in [0,1,2,3,4]:
            if int(doing) ==  1:
                card = './개인용카드.xlsx'
                wb = openpyxl.load_workbook(filename=card, data_only = True)
                ws = wb.active

                # ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ기본정보ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                #성명 B5, 연락처 E5, 주민등록번호 B6, 우편번호 B7, 주소 B8, 이메일 E7

                print("기본 정보 사항을 입력합니다. \n ")
                name = input("성함 : ")
                tel = input("연락처 : ")
                num = input("주민등록번호 : ")
                pnum = input("우편번호 : ")
                ad = input("주소 : ")
                email = input("E-mail : ")

                ws['B5'] = name
                ws['E5'] = tel
                ws['B6'] = num
                ws['B7'] = pnum
                ws['B8'] = ad
                ws['E7'] = email

                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ직장정보ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                #직장명 B10, 직장주소 B11, 사업자번호  E10, 연락처 B12, 팩스 E12

                print("\n직장 정보 사항이 있습니까? \n")
                while True:
                    cnum = input("네 : Y\n아니오 : N")
                    if cnum.upper() == 'N':
                        break
                    elif cnum.upper() == 'Y':
                        com_name = input("직장명 : ")
                        com_ad = input("직장 주소 : ")
                        com_tel = input("직장 연락처 : ")
                        com_num = input("사업자 번호 : ")
                        Fax = input("Fax : ")
                        ws['B10'] = com_name
                        ws['B11'] = com_ad
                        ws['E10'] = com_tel
                        ws['B12'] = com_num
                        ws['E12'] = Fax
                        break
                    else:
                        print("문자를 잘못 입력하셨습니다. 맞으면 Y, 아니면 N을 입력하세요\n")

                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ가족관계정보ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ

                print("\n가족관계사항이 있습니까? \n")
                while True:
                    cnum = input("네 : Y\n아니오 : N")
                    if cnum.upper() == 'N':
                        break
                    elif cnum.upper() == 'Y':
                        while True :
                            try :
                                n_f = int(input("몇 명을 추가하시겠습니까? : ")) #정수만 입력받아야 함, try 추가, 아래 코드들도 수정
                                while True:
                                    if type(n_f) == int:     
                                        loopnum = 0
                                        while True:
                                            if n_f - loopnum == 0:
                                                break
                                            else:
                                                fam_rel = input("관계 :")
                                                fam_name = input("성명 :")
                                                fam_num = input("주민번호 :")
                                                fam_tel = input("연락처 :")
                                                if loopnum == 0:
                                                    ws['H6'] = fam_rel
                                                    ws['I6'] = fam_name
                                                    ws['J6'] = fam_num
                                                    ws['K6'] = fam_tel
                                                    loopnum += 1
                                                elif loopnum == 1:
                                                    ws['H7'] = fam_rel
                                                    ws['I7'] = fam_name
                                                    ws['J7'] = fam_num
                                                    ws['K7'] = fam_tel
                                                    loopnum += 1
                                                elif loopnum == 2:
                                                    ws['H8'] = fam_rel
                                                    ws['I8'] = fam_name
                                                    ws['J8'] = fam_num
                                                    ws['K8'] = fam_tel
                                                    loopnum += 1
                                                elif loopnum == 3:
                                                    ws['H9'] = fam_rel
                                                    ws['I9'] = fam_name
                                                    ws['J9'] = fam_num
                                                    ws['K9'] = fam_tel
                                                    loopnum += 1
                                                elif loopnum == 4:
                                                    ws['H10'] = fam_rel
                                                    ws['I10'] = fam_name
                                                    ws['J10'] = fam_num
                                                    ws['K10'] = fam_tel
                                                    loopnum += 1
                                                elif loopnum == 5:
                                                    ws['H11'] = fam_rel
                                                    ws['I11'] = fam_name
                                                    ws['J11'] = fam_num
                                                    ws['K11'] = fam_tel
                                                    loopnum += 1
                                                elif loopnum == 6:
                                                    ws['H12'] = fam_rel
                                                    ws['I12'] = fam_name
                                                    ws['J12'] = fam_num
                                                    ws['K12'] = fam_tel
                                                    loopnum += 1                       
                                        break
                                break
                            except ValueError :
                                print("잘못 입력하셨습니다. 정수를 입력해주세요.\n")
                        break
                    else:
                        print("문자를 잘못 입력하셨습니다. 맞으면 Y, 아니면 N을 입력하세요\n")     

                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ자동차 보유현황ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                print("\n자동차 보유현황이 있습니까? \n")
                while True:
                    cnum = input("네 : Y\n아니오 : N")
                    if cnum.upper() == 'N':
                        break
                    elif cnum.upper() == 'Y':
                        while True:
                            try :
                                n_c = int(input("몇 대를 추가하시겠습니까? : "))
                                while True:
                                    if type(n_c) == int:
                                        loopnum = 0
                                        while True:
                                            if n_c - loopnum == 0:
                                                break
                                            else:
                                                car_num = input("차량번호 :")
                                                car_com = input("가입회사 :")
                                                car_end = input("만기일자 :")
                                                car_name = input("차명 :")
                                                if loopnum == 0:
                                                    ws['A15'] = car_num
                                                    ws['C15'] = car_com
                                                    ws['D15'] = car_end
                                                    ws['F15'] = car_name
                                                    loopnum += 1
                                                elif loopnum == 1:
                                                    ws['A16'] = car_num
                                                    ws['C16'] = car_com
                                                    ws['D16'] = car_end
                                                    ws['F16'] = car_name
                                                    loopnum += 1
                                                elif loopnum == 2:
                                                    ws['A17'] = car_num
                                                    ws['C17'] = car_com
                                                    ws['D17'] = car_end
                                                    ws['F17'] = car_name
                                                    loopnum += 1
                                                elif loopnum == 3:
                                                    ws['A18'] = car_num
                                                    ws['C18'] = car_com
                                                    ws['D18'] = car_end
                                                    ws['F18'] = car_name
                                                    loopnum += 1
                                                elif loopnum == 4:
                                                    ws['A19'] = car_num
                                                    ws['C19'] = car_com
                                                    ws['D19'] = car_end
                                                    ws['F19'] = car_name
                                                    loopnum += 1
                                                elif loopnum == 5:
                                                    ws['A20'] = car_num
                                                    ws['C20'] = car_com
                                                    ws['D20'] = car_end
                                                    ws['F20'] = car_name
                                                    loopnum += 1
                                        break
                                break
                            except ValueError:
                                print("잘못 입력하셨습니다. 정수를 다시 입력해주세요.\n")
                        break
                    else:
                        print("문자를 잘못 입력하셨습니다. 맞으면 Y, 아니면 N을 입력하세요\n")

                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ자사보험 가입현황 및 상태ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ

                print("\n자사보험 가입현황 및 상태가 있습니까? \n")
                while True:
                    cnum = input("네 : Y\n아니오 : N")
                    if cnum.upper() == 'N':
                        break
                    elif cnum.upper() == 'Y':
                        while True :
                            try :
                                n_i = int(input("몇 개를 추가하시겠습니까? : "))
                                while True:
                                    if type(n_i) == int:        
                                        loopnum = 0
                                        while True:
                                            if n_i - loopnum == 0:
                                                break
                                            else:
                                                contractor = input("계약자 :")
                                                client = input("피보험자 :")
                                                item = input("상품명 :")
                                                iduration = input("보험기간 :")
                                                con_num = input("계약/청약 번호 :")
                                                ins_fee = input("보험료 :")
                                                condition = input("상태 : ")
                                                if loopnum == 0:
                                                    ws['A23'] = contractor
                                                    ws['B23'] = client
                                                    ws['C23'] = item
                                                    ws['E23'] = iduration
                                                    ws['H23'] = con_num
                                                    ws['J23'] = ins_fee
                                                    ws['K23'] = condition
                                                    loopnum += 1
                                                elif loopnum == 1:
                                                    ws['A24'] = contractor
                                                    ws['B24'] = client
                                                    ws['C24'] = item
                                                    ws['E24'] = iduration
                                                    ws['H24'] = con_num
                                                    ws['J24'] = ins_fee
                                                    ws['K24'] = condition
                                                    loopnum += 1
                                                elif loopnum == 2:
                                                    ws['A25'] = contractor
                                                    ws['B25'] = client
                                                    ws['C25'] = item
                                                    ws['E25'] = iduration
                                                    ws['H25'] = con_num
                                                    ws['J25'] = ins_fee
                                                    ws['K25'] = condition
                                                    loopnum += 1
                                                elif loopnum == 3:
                                                    ws['A26'] = contractor
                                                    ws['B26'] = client
                                                    ws['C26'] = item
                                                    ws['E26'] = iduration
                                                    ws['H26'] = con_num
                                                    ws['J26'] = ins_fee
                                                    ws['K26'] = condition
                                                    loopnum += 1
                                                elif loopnum == 4:
                                                    ws['A27'] = contractor
                                                    ws['B27'] = client
                                                    ws['C27'] = item
                                                    ws['E27'] = iduration
                                                    ws['H27'] = con_num
                                                    ws['J27'] = ins_fee
                                                    ws['K27'] = condition
                                                    loopnum += 1
                                                elif loopnum == 5:
                                                    ws['A28'] = contractor
                                                    ws['B28'] = client
                                                    ws['C28'] = item
                                                    ws['E28'] = iduration
                                                    ws['H28'] = con_num
                                                    ws['J28'] = ins_fee
                                                    ws['K28'] = condition
                                                    loopnum += 1
                                                elif loopnum == 6:
                                                    ws['A29'] = contractor
                                                    ws['B29'] = client
                                                    ws['C29'] = item
                                                    ws['E29'] = iduration
                                                    ws['H29'] = con_num
                                                    ws['J29'] = ins_fee
                                                    ws['K29'] = condition
                                                    loopnum += 1
                                        break
                                break
                            except ValueError:
                                print("숫자를 잘못 입력하셨습니다. 정수를 입력해주세요.")
                    else:
                        print("문자를 잘못 입력하셨습니다. 맞으면 Y, 아니면 N을 입력하세요\n") 

                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ특이사항 및 참고사항ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ

                print("\n특이사항 및 참고사항이 있습니까? \n")
                while True:
                    cnum = input("네 : Y\n아니오 : N")
                    if cnum.upper() == 'N':
                        break
                    elif cnum.upper() == 'Y':
                        while True:
                            try : 
                                n_s = int(input("몇 개를 추가하시겠습니까? : "))
                                while True:
                                    if type(n_s) == int:
                                        loopnum = 0
                                        while True:
                                            if n_s - loopnum == 0:
                                                break
                                            else:
                                                specialcheck = input("특이사항 및 참고사항 :")
                                                if loopnum == 0:
                                                    ws['A32'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 1:
                                                    ws['A33'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 2:
                                                    ws['A34'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 3:
                                                    ws['A35'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 4:
                                                    ws['A36'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 5:
                                                    ws['A37'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 6:
                                                    ws['A38'] = specialcheck
                                                    loopnum += 1
                                                elif loopnum == 7:
                                                    ws['A39'] = specialcheck
                                                    loopnum += 1
                                        break
                                break
                            except ValueError:
                                print("잘못 입력하셨습니다. 정수를 다시 입력해주세요.\n")
                        break
                    else:
                        print("int(문자를 잘못 입력하셨습니다. 맞으면 Y, 아니면 N을 입력하세요\n")

                print("입력이 모두 완료 되었습니다.\n")
                print(name + '.xlsx 의 엑셀파일로 저장합니다. \n')
                wb.save(name + '.xlsx')
                print("저장이 완료 되었습니다.")
            elif int(doing) == 2:
                folder = './'
                real_cus_names = os.listdir(folder)

                checklist5 = [0,1,2,3,4,5]
                checklist6 = [0,1,2,3,4,5,6]
                checklist7 = [0,1,2,3,4,5,6,7]
                checklist8 = [0,1,2,3,4,5,6,7,8]

                while True:
                    input_cus_name = input("정보를 수정할 고객의 이름을 입력하세요\n \"종료\"를 입력하면 종료합니다.\n")
                    if input_cus_name == '종료':
                        break
                    elif input_cus_name + '.xlsx' in real_cus_names:
                        revise_cus = './' + input_cus_name + '.xlsx'
                        wb = openpyxl.load_workbook(filename = revise_cus, data_only = True)
                        ws = wb.active
                        while True:
                            if input_cus_name == 0:
                                break
                            else:
                                while True:
                                    revise_info_num = int(input("1 : 기본정보사항 수정 \n 2 : 직장정보사항 수정 \n 3 : 가족관계사항 수정 \n 4 : 자동차 보유현황 수정 \n 5 : 자사보험 가입현황 및 상태 수정\n 6. 특이사항 및 참고사항 수정\n 0. 종료\n"))
                                    if revise_info_num == 0:
                                        break
                                    elif revise_info_num == 1:
                                        standard(input_cus_name)
                                        while True:
                                            basic_num = int(input("1 : 연락처 수정 \n 2 : 우편번호 수정 \n 3 : 주소 수정 \n 4 : 이메일 수정 \n 5 : 이름 수정\n 0 : 종료\n"))
                                            if basic_num in checklist5:
                                                break
                                            else:
                                                print("0부터 5사이의 정수를 입력하세요.")
                                        if basic_num == 0:
                                            break
                                        elif basic_num == 1:
                                            rev_pnum = input("연락처를 입력하세요. :")
                                            ws['E5'] = rev_pnum
                                            standard(input_cus_name)
                                        elif basic_num == 2:
                                            rev_adnum = input("우편번호를 입력하세요. :")
                                            ws['B7'] = rev_adnum
                                            standard(input_cus_name)
                                        elif basic_num == 3:
                                            rev_ad = input("주소를 입력하세요. :")
                                            ws['B8'] = rev_ad
                                            standard(input_cus_name)
                                        elif basic_num == 4:
                                            rev_email = input("이메일을 입력하세요. :")
                                            ws['E7'] = rev_email
                                            standard(input_cus_name)
                                        elif basic_num == 5:
                                            rev_name = input("이름을 입력하세요. :")
                                            ws['B5'] = rev_name
                                            standard(input_cus_name)
                # ㅡ        ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ직장정보ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                                    elif revise_info_num == 2:
                                        company(input_cus_name)
                                        while True:
                                            comp_num = int(input("1 : 직장명 수정 \n 2 : 직장 주소 수정 \n 3 : 직장 연락처 수정 \n 4 : 사업자 번호 수정 \n 5 : Fax 수정\n 0 : 종료\n"))
                                            if comp_num in checklist5:
                                                break
                                            else:
                                                print("0부터 5사이의 정수를 입력하세요.")
                                        if comp_num == 0:
                                            break
                                        elif comp_num == 1:
                                            ws['B10'] = input("직장명을 입력하세요. :")
                                            company(input_cus_name)
                                        elif comp_num == 2:
                                            ws['B11'] = input("직장 주소를 입력하세요. :")
                                            company(input_cus_name)
                                        elif comp_num == 3:
                                            ws['E10'] = input("직장 연락처를 입력하세요. :")
                                            company(input_cus_name)
                                        elif comp_num == 4:
                                            ws['B12'] = input("사업자 번호를 입력하세요. :")
                                            company(input_cus_name)
                                        elif comp_num == 5:
                                            ws['E12'] = input("Fax를 입력하세요. :")
                                            company(input_cus_name)
                #ㅡ     ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ가족관계사항ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                                    elif revise_info_num == 3:
                                        family(input_cus_name)
                                        while True:
                                            family_num = int(input("1 : 첫번째 가족관계 수정 \n 2 : 두번째 가족관계 수정 \n 3 : 세번째 가족관계수정 \n 4 : 네번째 가족관계 수정 \n 5 : 다섯번째 가족관계 수정 \n 6 : 여섯번째 가족관계 수정 \n 7: 일곱번째 가족관계 수정\n 0 : 종료\n"))
                                            if family_num in checklist7:
                                                break
                                            else:
                                                print("0부터 7사이의 정수를 입력해주세요.")
                                        if family_num == 0:
                                            break
                                        elif family_num == 1:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H6'] = rev_fam_rel
                                            ws['I6'] = rev_fam_name
                                            ws['J6'] = rev_fam_num
                                            ws['K6'] = rev_fam_name
                                            family(input_cus_name)
                                        elif family_num == 2:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H7'] = rev_fam_rel
                                            ws['I7'] = rev_fam_name
                                            ws['J7'] = rev_fam_num
                                            ws['K7'] = rev_fam_name
                                            family(input_cus_name)
                                        elif family_num == 3:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H8'] = rev_fam_rel
                                            ws['I8'] = rev_fam_name
                                            ws['J8'] = rev_fam_num
                                            ws['K8'] = rev_fam_name
                                            family(input_cus_name)
                                        elif family_num == 4:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H9'] = rev_fam_rel
                                            ws['I9'] = rev_fam_name
                                            ws['J9'] = rev_fam_num
                                            ws['K9'] = rev_fam_name
                                            family(input_cus_name)
                                        elif family_num == 5:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H10'] = rev_fam_rel
                                            ws['I10'] = rev_fam_name
                                            ws['J10'] = rev_fam_num
                                            ws['K10'] = rev_fam_name
                                            family(input_cus_name)
                                        elif family_num == 6:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H11'] = rev_fam_rel
                                            ws['I11'] = rev_fam_name
                                            ws['J11'] = rev_fam_num
                                            ws['K11'] = rev_fam_name
                                            family(input_cus_name)
                                        elif family_num == 7:
                                            rev_fam_rel = input("가족관계 :")
                                            rev_fam_name = input("성명 :")
                                            rev_fam_num = input("주민등록번호 :")
                                            rev_fam_name = input("연락처 :")
                                            ws['H12'] = rev_fam_rel
                                            ws['I12'] = rev_fam_name
                                            ws['J12'] = rev_fam_num
                                            ws['K12'] = rev_fam_name
                                            family(input_cus_name)
                #ㅡ     ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ자동차정보ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                                    elif revise_info_num == 4:
                                        car(input_cus_name)
                                        while True:
                                            own_num = int(input("1 : 첫번째 자동차 수정 \n 2 : 두번째 자동차 수정 \n 3 : 세번째 자동차 수정 \n 4 : 네번째 자동차 수정 \n 5 : 다섯번째 자동차 수정 \n 6 : 여섯번째 자동차 수정 \n0 : 종료\n"))
                                            if own_num in checklist6:
                                                break
                                            else:
                                                print("0부터 6사이의 정수를 입력해주세요.")
                                        if own_num == 0:
                                            break
                                        elif own_num == 1:
                                            rev_car_num = input("차량번호 :")
                                            rev_car_com = input("가입회사 :")
                                            rev_car_end = input("만기일자 :")
                                            rev_car_name = input("차명 :")
                                            ws['A15'] = rev_car_num
                                            ws['C15'] = rev_car_com
                                            ws['D15'] = rev_car_end
                                            ws['F15'] = rev_car_name
                                            car(input_cus_name)
                                        elif own_num == 2:
                                            rev_car_num = input("차량번호 :")
                                            rev_car_com = input("가입회사 :")
                                            rev_car_end = input("만기일자 :")
                                            rev_car_name = input("차명 :")
                                            ws['A16'] = rev_car_num
                                            ws['C16'] = rev_car_com
                                            ws['D16'] = rev_car_end
                                            ws['F16'] = rev_car_name
                                            car(input_cus_name)
                                        elif own_num == 3:
                                            rev_car_num = input("차량번호 :")
                                            rev_car_com = input("가입회사 :")
                                            rev_car_end = input("만기일자 :")
                                            rev_car_name = input("차명 :")
                                            ws['A17'] = rev_car_num
                                            ws['C17'] = rev_car_com
                                            ws['D17'] = rev_car_end
                                            ws['F17'] = rev_car_name
                                            car(input_cus_name)
                                        elif own_num == 4:
                                            rev_car_num = input("차량번호 :")
                                            rev_car_com = input("가입회사 :")
                                            rev_car_end = input("만기일자 :")
                                            rev_car_name = input("차명 :")
                                            ws['A18'] = rev_car_num
                                            ws['C18'] = rev_car_com
                                            ws['D18'] = rev_car_end
                                            ws['F18'] = rev_car_name
                                            car(input_cus_name)
                                        elif own_num == 5:
                                            rev_car_num = input("차량번호 :")
                                            rev_car_com = input("가입회사 :")
                                            rev_car_end = input("만기일자 :")
                                            rev_car_name = input("차명 :")
                                            ws['A19'] = rev_car_num
                                            ws['C19'] = rev_car_com
                                            ws['D19'] = rev_car_end
                                            ws['F19'] = rev_car_name
                                            car(input_cus_name)
                                        elif own_num == 6:
                                            rev_car_num = input("차량번호 :")
                                            rev_car_com = input("가입회사 :")
                                            rev_car_end = input("만기일자 :")
                                            rev_car_name = input("차명 :")
                                            ws['A20'] = rev_car_num
                                            ws['C20'] = rev_car_com
                                            ws['D20'] = rev_car_end
                                            ws['F20'] = rev_car_name
                                            car(input_cus_name)
                #ㅡ     ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ자사보험 가입현황 및 상태ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                                    elif revise_info_num == 5:
                                        insurance(input_cus_name)
                                        while True:
                                            ins_num = int(input("1 : 첫번째 보험 수정 \n 2 : 두번째 보험 수정 \n 3 : 세번째 보험 수정 \n 4 : 네번째 보험 수정 \n 5 : 다섯번째 보험 수정 \n 6 : 여섯번째 보험 수정 \n 7 : 일곱번째 보험 수정 \n0 : 종료\n"))
                                            if ins_num in checklist7:
                                                break
                                            else:
                                                print("0부터 7사이의 정수를 입력해주세요.")
                                        if ins_num == 0:
                                            break
                                        elif ins_num == 1:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A23'] = rev_contractor
                                            ws['B23'] = rev_client
                                            ws['C23'] = rev_item
                                            ws['E23'] = rev_iduration
                                            ws['H23'] = rev_con_num
                                            ws['J23'] = rev_ins_fee
                                            ws['K23'] = rev_condition
                                            insurance(input_cus_name)   
                                        elif ins_num == 2:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A24'] = rev_contractor
                                            ws['B24'] = rev_client
                                            ws['C24'] = rev_item
                                            ws['E24'] = rev_iduration
                                            ws['H24'] = rev_con_num
                                            ws['J24'] = rev_ins_fee
                                            ws['K24'] = rev_condition
                                            insurance(input_cus_name)
                                        elif ins_num == 3:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A25'] = rev_contractor
                                            ws['B25'] = rev_client
                                            ws['C25'] = rev_item
                                            ws['E25'] = rev_iduration
                                            ws['H25'] = rev_con_num
                                            ws['J25'] = rev_ins_fee
                                            ws['K25'] = rev_condition
                                            insurance(input_cus_name)
                                        elif ins_num == 4:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A26'] = rev_contractor
                                            ws['B26'] = rev_client
                                            ws['C26'] = rev_item
                                            ws['E26'] = rev_iduration
                                            ws['H26'] = rev_con_num
                                            ws['J26'] = rev_ins_fee
                                            ws['K26'] = rev_condition
                                            insurance(input_cus_name)
                                        elif ins_num == 5:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A27'] = rev_contractor
                                            ws['B27'] = rev_client
                                            ws['C27'] = rev_item
                                            ws['E27'] = rev_iduration
                                            ws['H27'] = rev_con_num
                                            ws['J27'] = rev_ins_fee
                                            ws['K27'] = rev_condition
                                            insurance(input_cus_name)   
                                        elif ins_num == 6:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A28'] = rev_contractor
                                            ws['B28'] = rev_client
                                            ws['C28'] = rev_item
                                            ws['E28'] = rev_iduration
                                            ws['H28'] = rev_con_num
                                            ws['J28'] = rev_ins_fee
                                            ws['K28'] = rev_condition
                                            insurance(input_cus_name)
                                        elif ins_num == 7:
                                            rev_contractor = input("계약자 :")
                                            rev_client = input("피보험자 :")
                                            rev_item = input("상품명 :")
                                            rev_iduration = input("보험기간 :")
                                            rev_con_num = input("계약/청약번호")
                                            rev_ins_fee = input("보험료 :")
                                            rev_condition = input("상태 :")
                                            ws['A29'] = rev_contractor
                                            ws['B29'] = rev_client
                                            ws['C29'] = rev_item
                                            ws['E29'] = rev_iduration
                                            ws['H29'] = rev_con_num
                                            ws['J29'] = rev_ins_fee
                                            ws['K29'] = rev_condition
                                            insurance(input_cus_name)
                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ특이사항 및 참고사항ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ    
                                    elif revise_info_num == 6:
                                        special(input_cus_name)
                                        while True:
                                            spe_num = int(input("1 : 첫번째 특이,참고사항 수정 \n 2 : 두번째 특이,참고사항 수정 \n 3 : 세번째 특이,참고사항 수정 \n 4 : 네번째 특이,참고사항 수정 \n 5 : 다섯번째 특이,참고사항 수정 \n 6 : 여섯번째 특이,참고사항 수정 \n 7 : 일곱번째 특이,참고사항 수정 \n 8 : 여덟번째 특이,참고사항 수정\n 0 : 종료\n"))
                                            if spe_num in checklist8:
                                                break
                                            else:
                                                print("0부터 8사이의 정수를 입력해주세요.")
                                        if spe_num == 0:
                                            break
                                        elif spe_num == 1:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A32'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 2:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A33'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 3:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A34'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 4:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A35'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 5:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A36'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 6:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A37'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 7:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A38'] = rev_special
                                            special(input_cus_name)
                                        elif spe_num == 8:
                                            rev_special = input("특이/참고사항 입력 :")
                                            ws['A39'] = rev_special
                                            special(input_cus_name)
                                break
                    else:
                        print("그 고객명은 리스트에 존재하지 않습니다.\n 고객명을 다시 입력해주세요.\n")

                #ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ수정 및 추가 시작ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ
                save_name = ws['B5'].value
                print("입력이 모두 완료 되었습니다. \n")
                print("수정된 사항을 반영하여 엑셀파일을 저장합니다. \n")
                try: #개명으로 파일명이 바뀔경우
                    wb.save(save_name + '1.xlsx')
                    os.remove(save_name + '.xlsx')
                    os.rename('./'+save_name+'1.xlsx', save_name+'.xlsx')
                except:
                    print("개명으로 인해 파일명을 " + save_name + ".xlsx 로 저장합니다.")
                    before_name = input("개명전의 이름을 입력해주세요.")
                    os.remove(before_name + '.xlsx')
                    os.rename('./'+save_name+'1.xlsx', save_name + ".xlsx")
                print("저장이 완료 되었습니다. \n")
            elif int(doing) == 3:
                folder = './'
                real_cus_names = os.listdir(folder)

                while True:
                    input_cus_name = input("정보를 조회할 고객의 이름을 입력하세요. \"종료\"를 입력하면 종료합니다.\n")
                    if input_cus_name == '종료' :
                        break
                    elif input_cus_name + '.xlsx' in real_cus_names:
                        find_cus = './' + input_cus_name + '.xlsx'
                        wb = openpyxl.load_workbook(filename = find_cus, data_only = True)
                        ws = wb.active
                        find_info(input_cus_name)    
                        next = input("다음으로 넘어가려면 아무거나 입력하세요.")
                        break
            elif int(doing) == 4:
                folder = './'
                real_cus_names = os.listdir(folder)
                while True:
                    input_cus_name = input("정보를 삭제할 고객의 이름을 입력하세요. \"종료\"를 입력하면 종료합니다.\n")
                    if input_cus_name == '종료' :
                        break
                    elif input_cus_name + '.xlsx' in real_cus_names:
                        last_check = input("정말 삭제하시겠습니까? Y/N")
                        if last_check.upper() == 'Y':
                            os.remove(input_cus_name + '.xlsx')
                            print("삭제가 완료되었습니다.")
                            break
                        else:
                            print("메뉴로 돌아갑니다.")
                            break
            elif int(doing) == 0:
                print("프로그램을 이용해 주셔서 감사합니다.\n")
                print("프로그램을 종료합니다.")
                break
        else:
            print("잘못 입력하셨습니다. 다시 입력해주세요.\n")
    except ValueError:
        print("잘못 입력하셨습니다. 다시 입력해주세요.\n")
        
